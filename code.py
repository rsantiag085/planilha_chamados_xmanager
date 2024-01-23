import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Função para buscar chamados para um usuário
def buscar_chamados(idUsuario, nome, auth):
    url = f'https://helpdesk.seudominio.com.br/xmanager-services/service/xdeskAcompanhamento/chamados?idUsuario={idUsuario}'
    response = requests.get(url, auth=auth)
    return response.json()['chamados']

# Define informações de autenticação
auth = HTTPBasicAuth('usuário', 'senha')

# Lista de usuários e seus IDs
usuarios = [
    {'nome': 'usuario01', 'idUsuario': xxxx},
    {'nome': 'usuario02', 'idUsuario': xxxx},
    {'nome': 'usuario03', 'idUsuario': xxxx},
    {'nome': 'usuario04', 'idUsuario': xxxx}
]

# Criar uma lista para armazenar os dados dos chamados
chamados_data = []

# Data atual
data_atual = date.today()
data_atual_str = data_atual.strftime('%Y-%m-%d')

# Função para ajustar o status
def ajustar_status(status):
    if status == 'AGUARDANDO RESPOSTA DO SOLICITANTE':
        return 'AGUARDANDO SOLICITANTE'
    elif status == 'AGUARDANDO RETORNO DE FORNECEDOR':
        return 'AGUARDANDO FORNECEDOR'
    return status


# Loop através dos usuários
for usuario in usuarios:
    nome = usuario['nome']
    chamados = buscar_chamados(usuario['idUsuario'], nome, auth)

    for chamado in chamados:
        if (nome in chamado['detalhes'][2]['value'] and chamado['status'] not in ['ENCERRADO', 'CANCELADO']) or (date.today().strftime('%d/%m/%Y') in chamado['dtUltimaMovimentacao'] and chamado['status'] == 'ENCERRADO' and nome in chamado['detalhes'][2]['value']):
            chamado_data = {
                'Número do chamado': chamado['id'],
                'Assunto': chamado['assunto'],
                'Status': ajustar_status(chamado['status']),  # Use a função para ajustar o status aqui
            }
            chamados_data.append(chamado_data)

# Criar um DataFrame com os dados dos chamados
df = pd.DataFrame(chamados_data)

# Construir o nome do arquivo com o caminho completo para C:\relatórios
nome_arquivo = f'C:/Users/seu_usuario/Meu Drive/relatórios/chamados_{data_atual_str}.xlsx'

# Cria um novo Workbook
wb = Workbook()

# Seleciona a planilha padrão (Sheet)
ws = wb.active

# Define a fonte CALIBRI tamanho 12 e em negrito
font = Font(name='CALIBRI', size=12, bold=True)

# Define o alinhamento centralizado para todas as células
alignment = Alignment(horizontal='center', vertical='center')

# Aplica a fonte e o alinhamento às células do cabeçalho
for col_num, col in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_num, value=col)
    cell.font = font
    cell.alignment = alignment

# Preenche os dados
for row_num, (_, row) in enumerate(df.iterrows(), 2):
    for col_num, value in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        cell.font = font
        cell.alignment = alignment

# Salva o arquivo XLS
wb.save(nome_arquivo)

print(f'Chamados salvos em {nome_arquivo}')
